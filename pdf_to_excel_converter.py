from pdfminer.high_level import extract_text
from openpyxl import Workbook
import re

def extract_tables_from_pdf(pdf_path):
    """
    Extract tables from a PDF file.

    Args:
    - pdf_path (str): Path to the PDF file.

    Returns:
    - list: List of tables, where each table is a list of rows containing cell values.
    """
    text = extract_text(pdf_path)
    tables = extract_tables_from_text(text)
    return tables

def extract_tables_from_text(text):
    """
    Extract tables from text using regex patterns.

    Args:
    - text (str): Text extracted from a PDF.

    Returns:
    - list: List of tables, where each table is a list of rows containing cell values.
    """
    tables = []
    table_pattern = re.compile(r'\+.*?\+[\n\S]*?(?:\|.*?\|[\n\S]*?)*\+')

    matches = table_pattern.findall(text)
    for match in matches:
        table_data = []
        rows = match.strip().split('\n')
        for row in rows:
            cells = [cell.strip() for cell in row.strip('|').split('|')]
            table_data.append(cells)
        tables.append(table_data)

    return tables

def write_to_excel(tables, output_file):
    """
    Write tables to an Excel file.

    Args:
    - tables (list): List of tables to write to Excel.
    - output_file (str): Output Excel file path.
    """
    wb = Workbook()
    for idx, table in enumerate(tables, start=1):
        ws = wb.create_sheet(f"Table_{idx}")

        for row_idx, row in enumerate(table, start=1):
            for col_idx, cell in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell)

    wb.save(output_file)
    print(f"Tables successfully written to {output_file}")
