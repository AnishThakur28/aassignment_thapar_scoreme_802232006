# -*- coding: utf-8 -*-
"""Untitled4.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1D-Ojjs2FNiYAkPRTvN9TTqKK5nc5GAUi

### **Tool Detecting and Extracting Tables from PDFs** **bold text**

THIS is Developed by : ANISH THAKUR
"""

pip install pdfminer.six

"""## **By this code we can Automatically Download multiple pdf files to Excel files**"""

from pdfminer.high_level import extract_text
from openpyxl import Workbook
import re

def extract_tables_from_pdf(pdf_path):
    text = extract_text(pdf_path)
    tables = extract_tables_from_text(text)
    return tables

def extract_tables_from_text(text):
    tables = []
    table_pattern = re.compile(r'\+.*?\+[\n\S]*?(?:\|.*?\|[\n\S]*?)*\+')

    matches = table_pattern.findall(text)
    for match in matches:
        table_data = []
        rows = match.strip().split('\n')
        for row in rows:
            cells = [cell.strip() for cell in row.strip('|').split('|')]
            table_data.append(cells)
        tables.append(table_data)

    return tables

def write_to_excel(tables, output_file):
    wb = Workbook()

    for idx, table in enumerate(tables, start=1):
        ws = wb.create_sheet(f"Table_{idx}")

        for row_idx, row in enumerate(table, start=1):
            for col_idx, cell in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell)

    wb.save(output_file)
    print(f"Tables successfully written to {output_file}")

if __name__ == "__main__":
    pdf_path = r"/content/test3.pdf"
    pdf_path = r"/content/test5.pdf"
    pdf_path = r"/content/test6.pdf"
    output_excel = 'output.xlsx'

    tables = extract_tables_from_pdf(pdf_path)
    write_to_excel(tables, output_excel)

"""## **Guide to Detect and Extract Tables from PDFs**

1.   Step 1 : Run The Script Below
2.   Step 2 : Upload/Cancel the Pdf from Your Computer
3.   Step 3 : Just Chill and Wait.
4.   Step 4 : Excel File Automatically Downloaded


"""

from pdfminer.high_level import extract_text
from openpyxl import Workbook
import re
from google.colab import files

def extract_tables_from_pdf(pdf_path):

    text = extract_text(pdf_path)  # Extract text from PDF
    tables = extract_tables_from_text(text)  # Extract tables from extracted text
    return tables

def extract_tables_from_text(text):                                                         # Extracts tables from text using regex patterns.      Returns: list: List of tables, where each table is a list of rows containing cell values.
    tables = []
    table_pattern = re.compile(r'\+.*?\+[\n\S]*?(?:\|.*?\|[\n\S]*?)*\+')                    # Regular expression pattern to identify tables in text
    matches = table_pattern.findall(text)                                                   # Find all matches of the table pattern in the text

    for match in matches:                                                                   # Process each matched table
        table_data = []
        rows = match.strip().split('\n')                                                    # Split the matched table into rows
        for row in rows:                                                                    # Process each row to extract cells

            cells = [cell.strip() for cell in row.strip('|').split('|')]                    # Split row into cells based on '|' separator, strip whitespace
            table_data.append(cells)
        tables.append(table_data)

    return tables

def write_to_excel(tables, output_file):                                                     # Writes tables data to an Excel file using openpyxl.
    wb = Workbook()                                                                          # Create a new Workbook object
    for idx, table in enumerate(tables, start=1):                                            # Iterate through each table and write to a new sheet in the Workbook
        ws = wb.create_sheet(f"Table_{idx}")  # Create a new sheet for each table
        for row_idx, row in enumerate(table, start=1):                                       # Iterate through each row in the table

            for col_idx, cell in enumerate(row, start=1):                                    # Iterate through each cell in the row and write to the sheet
                ws.cell(row=row_idx, column=col_idx, value=cell)                             # Write cell value to sheet

    wb.save(output_file)                                                                     # Save the Workbook to the specified output file
    print(f"Tables successfully written to {output_file}")

def convert_pdf_to_excel_and_download():                                                     # Main function to convert uploaded PDF files to Excel and initiate download "

    uploaded = files.upload()
    pdf_paths = list(uploaded.keys())

    if not pdf_paths:                                                                         # Check if any PDF files were uploaded
        print("No PDF files uploaded.")
        return

    output_file = 'output.xlsx'                                                                # Output file name for the Excel file
    tables = []                                                                                # Convert PDFs to Excel tables and write to output file
    for pdf_path in pdf_paths:
        extracted_tables = extract_tables_from_pdf(pdf_path)                                   # Extract tables from each PDF
        tables.extend(extracted_tables)                                                        # Extend tables list with tables from current PDF

    write_to_excel(tables, output_file)                                                        # Write all tables to Excel fil
    files.download(output_file)                                                                # Download the generated Excel file to the user's local system
convert_pdf_to_excel_and_download()                                                            # Call the main function to start the PDF to Excel conversion and download process